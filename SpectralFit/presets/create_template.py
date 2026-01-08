"""
Script to generate material_presets.xlsx template with example materials.

Run this script to create the Excel template:
    python create_template.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_material_sheet(wb, sheet_name, settings_data, peak_data):
    """
    Create a material sheet with settings and peak templates.

    Parameters
    ----------
    wb : Workbook
        openpyxl workbook
    sheet_name : str
        Name for the sheet (e.g., 'Graphene_Raman')
    settings_data : dict
        Processing settings key-value pairs
    peak_data : list of dict
        Peak templates (one dict per peak)
    """
    ws = wb.create_sheet(sheet_name)

    # Row 1: Processing Settings Headers
    settings_headers = [
        'x_range_enabled', 'x_min', 'x_max', 'despike_threshold',
        'baseline_algorithm', 'baseline_degree', 'baseline_lambda',
        'baseline_p', 'description'
    ]

    for col_idx, header in enumerate(settings_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Row 2: Processing Settings Values
    for col_idx, header in enumerate(settings_headers, start=1):
        value = settings_data.get(header, '')
        ws.cell(row=2, column=col_idx, value=value)

    # Row 3: Blank separator

    # Row 4: Peak Template Headers
    peak_headers = [
        'peak_label', 'center', 'center_tolerance',
        'amplitude', 'width_fwhm', 'shape', 'color'
    ]

    for col_idx, header in enumerate(peak_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Row 5+: Peak Template Data
    for peak_idx, peak in enumerate(peak_data, start=5):
        for col_idx, header in enumerate(peak_headers, start=1):
            value = peak.get(header, '')
            ws.cell(row=peak_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, max(len(settings_headers), len(peak_headers)) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18


def create_presets_excel():
    """Create material_presets.xlsx with example materials."""

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ========== Graphene_Raman ==========
    graphene_settings = {
        'x_range_enabled': True,
        'x_min': 1200,
        'x_max': 2800,
        'despike_threshold': 6.0,
        'baseline_algorithm': 'ALS',
        'baseline_degree': '',
        'baseline_lambda': 10000,
        'baseline_p': 0.001,
        'description': 'Graphene on Si/SiO2 substrate, D/G/2D analysis'
    }

    graphene_peaks = [
        {
            'peak_label': 'D-band',
            'center': 1350,
            'center_tolerance': 20,
            'amplitude': 5000,
            'width_fwhm': 50,
            'shape': 0.5,
            'color': '#1f77b4'
        },
        {
            'peak_label': 'G-band',
            'center': 1580,
            'center_tolerance': 10,
            'amplitude': 8000,
            'width_fwhm': 60,
            'shape': 0.5,
            'color': '#ff7f0e'
        },
        {
            'peak_label': '2D-band',
            'center': 2700,
            'center_tolerance': 50,
            'amplitude': 6000,
            'width_fwhm': 80,
            'shape': 0.5,
            'color': '#2ca02c'
        }
    ]

    create_material_sheet(wb, 'Graphene_Raman', graphene_settings, graphene_peaks)

    # ========== MoS2_Raman ==========
    mos2_settings = {
        'x_range_enabled': False,
        'x_min': '',
        'x_max': '',
        'despike_threshold': 8.0,
        'baseline_algorithm': 'ALS',
        'baseline_degree': '',
        'baseline_lambda': 50000,
        'baseline_p': 0.01,
        'description': 'MoS2 E2g and A1g peaks'
    }

    mos2_peaks = [
        {
            'peak_label': 'E2g',
            'center': 383,
            'center_tolerance': 5,
            'amplitude': 10000,
            'width_fwhm': 10,
            'shape': 0.3,
            'color': '#d62728'
        },
        {
            'peak_label': 'A1g',
            'center': 408,
            'center_tolerance': 5,
            'amplitude': 12000,
            'width_fwhm': 12,
            'shape': 0.3,
            'color': '#9467bd'
        }
    ]

    create_material_sheet(wb, 'MoS2_Raman', mos2_settings, mos2_peaks)

    # ========== Silicon_Raman ==========
    silicon_settings = {
        'x_range_enabled': False,
        'x_min': '',
        'x_max': '',
        'despike_threshold': 6.0,
        'baseline_algorithm': 'Polynomial',
        'baseline_degree': 5,
        'baseline_lambda': '',
        'baseline_p': '',
        'description': 'Silicon reference peak at 520 cm⁻¹'
    }

    silicon_peaks = [
        {
            'peak_label': 'Si',
            'center': 520,
            'center_tolerance': 3,
            'amplitude': 15000,
            'width_fwhm': 8,
            'shape': 0.2,
            'color': '#2ca02c'
        }
    ]

    create_material_sheet(wb, 'Silicon_Raman', silicon_settings, silicon_peaks)

    # Save workbook
    output_path = 'material_presets.xlsx'
    wb.save(output_path)
    print(f"[OK] Created: {output_path}")
    print(f"   - Graphene_Raman: 3 peaks (D, G, 2D)")
    print(f"   - MoS2_Raman: 2 peaks (E2g, A1g)")
    print(f"   - Silicon_Raman: 1 peak (Si)")
    print(f"\nTo add a new material:")
    print(f"1. Open {output_path} in Excel")
    print(f"2. Right-click sheet tab -> Insert -> New Sheet")
    print(f"3. Name it: MaterialName_Mode (e.g., 'GaN_Raman')")
    print(f"4. Copy structure from existing sheet")
    print(f"5. Edit settings (Row 2) and peaks (Row 5+)")
    print(f"6. Save and reload in SpectralFit app!")


if __name__ == '__main__':
    create_presets_excel()
